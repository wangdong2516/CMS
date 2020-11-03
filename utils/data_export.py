"""
    导出数据工具类的封装
"""
import openpyxl
from typing import List
from openpyxl.styles import Font
from openpyxl import load_workbook


class ExportXls(object):

    def __init__(self, filename):
        """
            初始化

            包括一些样式的设置(单元格的大小，自动居中，宽度，高度的设置)
        """
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.font = Font(name='Monaco', size=14, color='FF000000')
        if filename.endswith('.xls'):
            raise Exception("Don't Suport .xls Filename, Please Use Other Tools")
        self.filename = filename

    def add_sheet(self, sheet_name: str) -> None:
        """
            添加sheet
        :param sheet_name: sheet名称
        :return:
        """
        self.workbook.create_sheet(title=sheet_name)

    def write_cell_title(self, cell_title_list: List) -> None:
        """
            添加单元格表头内容
        :param cell_title:
        :return:
        """
        if isinstance(cell_title_list, list):
            self.sheet.append(cell_title_list)

    def write(self, data_list: List[List], /) -> None:
        """
            写入数据到xlsx文件
        :param data_list: 数据列表，列表嵌套的形式
        :return:
        """
        if isinstance(data_list, list):
            for data in data_list:
                self.sheet.append(data)
        else:
            raise Exception('An iterable object need')
        self.workbook.save(filename=self.filename)

    def read_title(self) -> List:
        """
            读取单元格表头内容
        :return: List: 表头数据列表
        """
        wb = load_workbook(filename=self.filename)
        for i in wb:
            for data in i.values:
                return list(data)

    def read(self):
        """
            读取xlsx文件的内容
        :return:
        """
        result = []
        wb = load_workbook(filename=self.filename)
        for i in wb:
            for data in i.values:
                result.append(list(data))
        return result[1:]


